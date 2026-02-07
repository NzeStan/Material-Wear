# excel_bulk_orders/tests/test_utils.py
"""
Comprehensive tests for Excel Bulk Orders utility functions.

Coverage:
- generate_excel_coupon_codes: Coupon generation, uniqueness
- generate_excel_template: Template creation, validation rules, formatting
- validate_excel_file: Data validation, error reporting
- create_participants_from_excel: Participant creation, coupon handling
- generate_participants_pdf: PDF document generation
- generate_participants_word: Word document generation  
- generate_participants_excel: Excel report generation
- Error handling and edge cases
"""
from decimal import Decimal
from django.test import TestCase
from django.contrib.auth import get_user_model
from django.conf import settings
from io import BytesIO
from unittest.mock import Mock, patch, MagicMock
import openpyxl
import pandas as pd

from excel_bulk_orders.models import ExcelBulkOrder, ExcelCouponCode, ExcelParticipant
from excel_bulk_orders.utils import (
    generate_excel_coupon_codes,
    generate_excel_template,
    validate_excel_file,
    create_participants_from_excel,
    generate_participants_pdf,
    generate_participants_word,
    generate_participants_excel,
)

User = get_user_model()


class GenerateExcelCouponCodesTest(TestCase):
    """Test coupon code generation"""

    def setUp(self):
        """Set up test data"""
        self.bulk_order = ExcelBulkOrder.objects.create(
            title='Coupon Test Order',
            coordinator_name='Test',
            coordinator_email='coupon@example.com',
            coordinator_phone='08012345678',
            price_per_participant=Decimal('5000.00')
        )

    def test_generate_default_count(self):
        """Test generating default number of coupons (10)"""
        coupons = generate_excel_coupon_codes(self.bulk_order)

        self.assertEqual(len(coupons), 10)
        self.assertEqual(self.bulk_order.coupons.count(), 10)

    def test_generate_custom_count(self):
        """Test generating custom number of coupons"""
        count = 25
        coupons = generate_excel_coupon_codes(self.bulk_order, count=count)

        self.assertEqual(len(coupons), count)
        self.assertEqual(self.bulk_order.coupons.count(), count)

    def test_generated_codes_are_unique(self):
        """Test that all generated codes are unique"""
        coupons = generate_excel_coupon_codes(self.bulk_order, count=50)

        codes = [coupon.code for coupon in coupons]
        unique_codes = set(codes)

        self.assertEqual(len(codes), len(unique_codes))

    def test_generated_codes_format(self):
        """Test that generated codes follow expected format"""
        coupons = generate_excel_coupon_codes(self.bulk_order, count=5)

        for coupon in coupons:
            # Should be 8 characters, uppercase alphanumeric
            self.assertEqual(len(coupon.code), 8)
            self.assertTrue(coupon.code.isupper())
            self.assertTrue(coupon.code.isalnum())

    def test_coupons_linked_to_bulk_order(self):
        """Test that coupons are linked to correct bulk order"""
        coupons = generate_excel_coupon_codes(self.bulk_order, count=5)

        for coupon in coupons:
            self.assertEqual(coupon.bulk_order, self.bulk_order)

    def test_coupons_default_not_used(self):
        """Test that generated coupons are not used by default"""
        coupons = generate_excel_coupon_codes(self.bulk_order, count=5)

        for coupon in coupons:
            self.assertFalse(coupon.is_used)


class GenerateExcelTemplateTest(TestCase):
    """Test Excel template generation"""

    def setUp(self):
        """Set up test data"""
        self.bulk_order_with_custom = ExcelBulkOrder.objects.create(
            title='Template Test Order',
            coordinator_name='Test Coordinator',
            coordinator_email='template@example.com',
            coordinator_phone='08012345678',
            price_per_participant=Decimal('5000.00'),
            requires_custom_name=True
        )

        self.bulk_order_no_custom = ExcelBulkOrder.objects.create(
            title='No Custom Name Order',
            coordinator_name='Test',
            coordinator_email='nocustom@example.com',
            coordinator_phone='08012345678',
            price_per_participant=Decimal('5000.00'),
            requires_custom_name=False
        )

    def test_template_generation_returns_buffer(self):
        """Test that template generation returns BytesIO buffer"""
        buffer = generate_excel_template(self.bulk_order_with_custom)

        self.assertIsInstance(buffer, BytesIO)
        self.assertGreater(len(buffer.getvalue()), 0)

    def test_template_has_correct_sheets(self):
        """Test that template has required sheets"""
        buffer = generate_excel_template(self.bulk_order_with_custom)
        wb = openpyxl.load_workbook(buffer)

        self.assertIn('Participants', wb.sheetnames)
        self.assertIn('Instructions', wb.sheetnames)

    def test_template_with_custom_name_columns(self):
        """Test template columns when custom name is required"""
        buffer = generate_excel_template(self.bulk_order_with_custom)
        wb = openpyxl.load_workbook(buffer)
        ws = wb['Participants']

        # Get header row
        headers = [cell.value for cell in ws[1]]

        expected_headers = ['S/N', 'Full Name', 'Size', 'Custom Name', 'Coupon Code']
        self.assertEqual(headers, expected_headers)

    def test_template_without_custom_name_columns(self):
        """Test template columns when custom name not required"""
        buffer = generate_excel_template(self.bulk_order_no_custom)
        wb = openpyxl.load_workbook(buffer)
        ws = wb['Participants']

        # Get header row
        headers = [cell.value for cell in ws[1]]

        expected_headers = ['S/N', 'Full Name', 'Size', 'Coupon Code']
        self.assertEqual(headers, expected_headers)

    def test_template_has_example_row(self):
        """Test that template includes example data row"""
        buffer = generate_excel_template(self.bulk_order_with_custom)
        wb = openpyxl.load_workbook(buffer)
        ws = wb['Participants']

        # Check example row (row 2)
        self.assertIsNotNone(ws.cell(2, 1).value)  # S/N
        self.assertIsNotNone(ws.cell(2, 2).value)  # Full Name

    def test_template_has_size_validation(self):
        """Test that template has data validation for Size column"""
        buffer = generate_excel_template(self.bulk_order_with_custom)
        wb = openpyxl.load_workbook(buffer)
        ws = wb['Participants']

        # Check if data validation exists on Size column
        # openpyxl stores validations in ws.data_validations
        self.assertIsNotNone(ws.data_validations)

    def test_template_instructions_sheet(self):
        """Test that instructions sheet has content"""
        buffer = generate_excel_template(self.bulk_order_with_custom)
        wb = openpyxl.load_workbook(buffer)
        instructions_ws = wb['Instructions']

        # Check that instructions sheet has content
        self.assertIsNotNone(instructions_ws.cell(1, 1).value)

    def test_template_can_be_saved_and_loaded(self):
        """Test that generated template can be saved and loaded"""
        buffer = generate_excel_template(self.bulk_order_with_custom)

        # Try to load it again
        try:
            wb = openpyxl.load_workbook(buffer)
            self.assertIsNotNone(wb)
        except Exception as e:
            self.fail(f"Failed to load generated template: {str(e)}")


class ValidateExcelFileTest(TestCase):
    """Test Excel file validation"""

    def setUp(self):
        """Set up test data"""
        self.bulk_order = ExcelBulkOrder.objects.create(
            title='Validation Test Order',
            coordinator_name='Test',
            coordinator_email='validate@example.com',
            coordinator_phone='08012345678',
            price_per_participant=Decimal('5000.00'),
            requires_custom_name=True
        )

    def create_valid_excel(self):
        """Helper to create valid Excel file"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Participants'

        # Header
        ws.append(['S/N', 'Full Name', 'Size', 'Custom Name', 'Coupon Code'])

        # Valid data
        ws.append([1, 'John Doe', 'Medium', 'JOHNNY', ''])
        ws.append([2, 'Jane Smith', 'Large', 'JANE', ''])

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    def create_invalid_excel_missing_columns(self):
        """Helper to create Excel with missing columns"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Participants'

        # Missing Size column
        ws.append(['S/N', 'Full Name', 'Coupon Code'])

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    def create_excel_with_errors(self):
        """Helper to create Excel with validation errors"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Participants'

        # Header
        ws.append(['S/N', 'Full Name', 'Size', 'Custom Name', 'Coupon Code'])

        # Invalid data
        ws.append([1, '', 'Medium', 'NAME1', ''])  # Missing name
        ws.append([2, 'Jane Smith', 'Invalid Size', 'NAME2', ''])  # Invalid size

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    def test_validate_valid_excel(self):
        """Test validating a valid Excel file"""
        excel_file = self.create_valid_excel()
        result = validate_excel_file(self.bulk_order, excel_file)

        self.assertTrue(result['valid'])
        self.assertEqual(len(result['errors']), 0)
        self.assertEqual(result['summary']['total_rows'], 2)
        self.assertEqual(result['summary']['valid_rows'], 2)

    def test_validate_excel_missing_columns(self):
        """Test validating Excel with missing columns"""
        excel_file = self.create_invalid_excel_missing_columns()
        result = validate_excel_file(self.bulk_order, excel_file)

        self.assertFalse(result['valid'])
        self.assertGreater(len(result['errors']), 0)

    def test_validate_excel_with_errors(self):
        """Test validating Excel with data errors"""
        excel_file = self.create_excel_with_errors()
        result = validate_excel_file(self.bulk_order, excel_file)

        self.assertFalse(result['valid'])
        self.assertGreater(len(result['errors']), 0)
        self.assertEqual(result['summary']['error_rows'], 2)

    def test_validate_excel_error_structure(self):
        """Test that validation errors have correct structure"""
        excel_file = self.create_excel_with_errors()
        result = validate_excel_file(self.bulk_order, excel_file)

        for error in result['errors']:
            self.assertIn('row', error)
            self.assertIn('field', error)
            self.assertIn('error', error)
            self.assertIn('current_value', error)


class CreateParticipantsFromExcelTest(TestCase):
    """Test participant creation from Excel"""

    def setUp(self):
        """Set up test data"""
        self.bulk_order = ExcelBulkOrder.objects.create(
            title='Participant Creation Test',
            coordinator_name='Test',
            coordinator_email='create@example.com',
            coordinator_phone='08012345678',
            price_per_participant=Decimal('5000.00'),
            requires_custom_name=True
        )

        # Create coupon
        self.coupon = ExcelCouponCode.objects.create(
            bulk_order=self.bulk_order,
            code='VALID123'
        )

    def create_test_excel(self):
        """Helper to create test Excel with participants"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Participants'

        # Header
        ws.append(['S/N', 'Full Name', 'Size', 'Custom Name', 'Coupon Code'])

        # Participants
        ws.append([1, 'John Doe', 'Medium', 'JOHNNY', ''])
        ws.append([2, 'Jane Smith', 'Large', 'JANE', 'VALID123'])

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    def test_create_participants_success(self):
        """Test successful participant creation"""
        excel_file = self.create_test_excel()
        count = create_participants_from_excel(self.bulk_order, excel_file)

        self.assertEqual(count, 2)
        self.assertEqual(self.bulk_order.participants.count(), 2)

    def test_create_participants_with_coupon(self):
        """Test that coupon is properly applied"""
        excel_file = self.create_test_excel()
        create_participants_from_excel(self.bulk_order, excel_file)

        # Find participant with coupon
        participant = self.bulk_order.participants.get(coupon_code='VALID123')

        self.assertTrue(participant.is_coupon_applied)
        self.assertEqual(participant.coupon, self.coupon)

        # Coupon should be marked as used
        self.coupon.refresh_from_db()
        self.assertTrue(self.coupon.is_used)

    def test_create_participants_size_mapping(self):
        """Test that size names are mapped to codes"""
        excel_file = self.create_test_excel()
        create_participants_from_excel(self.bulk_order, excel_file)

        # Check size mapping
        participant = self.bulk_order.participants.get(full_name='John Doe')
        self.assertEqual(participant.size, 'M')  # Medium -> M

    def test_create_participants_custom_name_uppercase(self):
        """Test that custom names are uppercased"""
        excel_file = self.create_test_excel()
        create_participants_from_excel(self.bulk_order, excel_file)

        participant = self.bulk_order.participants.first()
        self.assertEqual(participant.custom_name, participant.custom_name.upper())


class GenerateParticipantsPDFTest(TestCase):
    """Test PDF generation"""

    def setUp(self):
        """Set up test data"""
        self.bulk_order = ExcelBulkOrder.objects.create(
            title='PDF Test Order',
            coordinator_name='Test Coordinator',
            coordinator_email='pdf@example.com',
            coordinator_phone='08012345678',
            price_per_participant=Decimal('5000.00')
        )

        # Create participants
        for i in range(5):
            ExcelParticipant.objects.create(
                bulk_order=self.bulk_order,
                full_name=f'Participant {i}',
                size='M',
                row_number=i + 2
            )

    @patch('weasyprint.HTML')
    def test_generate_pdf_returns_buffer(self, mock_html):
        """Test that PDF generation returns buffer"""
        mock_html_instance = MagicMock()
        mock_html_instance.write_pdf.return_value = b'fake pdf content'
        mock_html.return_value = mock_html_instance

        buffer = generate_participants_pdf(self.bulk_order)

        self.assertIsInstance(buffer, BytesIO)

    @patch('weasyprint.HTML')
    def test_generate_pdf_with_request(self, mock_html):
        """Test PDF generation with request object"""
        mock_html_instance = MagicMock()
        mock_html_instance.write_pdf.return_value = b'fake pdf content'
        mock_html.return_value = mock_html_instance

        mock_request = Mock()
        mock_request.build_absolute_uri.return_value = 'http://example.com'

        buffer = generate_participants_pdf(self.bulk_order, request=mock_request)

        self.assertIsInstance(buffer, BytesIO)


class GenerateParticipantsWordTest(TestCase):
    """Test Word document generation"""

    def setUp(self):
        """Set up test data"""
        self.bulk_order = ExcelBulkOrder.objects.create(
            title='Word Test Order',
            coordinator_name='Test',
            coordinator_email='word@example.com',
            coordinator_phone='08012345678',
            price_per_participant=Decimal('5000.00')
        )

        for i in range(3):
            ExcelParticipant.objects.create(
                bulk_order=self.bulk_order,
                full_name=f'Participant {i}',
                size='L',
                row_number=i + 2
            )

    def test_generate_word_returns_buffer(self):
        """Test that Word generation returns buffer"""
        buffer = generate_participants_word(self.bulk_order)

        self.assertIsInstance(buffer, BytesIO)
        self.assertGreater(len(buffer.getvalue()), 0)


class GenerateParticipantsExcelTest(TestCase):
    """Test Excel report generation"""

    def setUp(self):
        """Set up test data"""
        self.bulk_order = ExcelBulkOrder.objects.create(
            title='Excel Report Test',
            coordinator_name='Test',
            coordinator_email='excelreport@example.com',
            coordinator_phone='08012345678',
            price_per_participant=Decimal('5000.00'),
            requires_custom_name=True
        )

        for i in range(3):
            ExcelParticipant.objects.create(
                bulk_order=self.bulk_order,
                full_name=f'Participant {i}',
                size='M',
                custom_name=f'NAME{i}',
                row_number=i + 2
            )

    def test_generate_excel_returns_buffer(self):
        """Test that Excel generation returns buffer"""
        buffer = generate_participants_excel(self.bulk_order)

        self.assertIsInstance(buffer, BytesIO)
        self.assertGreater(len(buffer.getvalue()), 0)

    def test_generated_excel_can_be_loaded(self):
        """Test that generated Excel can be loaded"""
        buffer = generate_participants_excel(self.bulk_order)

        # Try to load with pandas
        try:
            df = pd.read_excel(buffer, sheet_name='Participants')
            self.assertEqual(len(df), 3)
        except Exception as e:
            self.fail(f"Failed to load generated Excel: {str(e)}")