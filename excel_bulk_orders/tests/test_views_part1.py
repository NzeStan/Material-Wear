# excel_bulk_orders/tests/test_views.py
"""
Comprehensive tests for Excel Bulk Orders API Views.

Coverage:
- ExcelBulkOrderViewSet: CRUD operations, permissions, queryset filtering
- Upload action: File upload, validation, Cloudinary integration
- Validate action: Excel validation, error reporting
- Initialize payment: Paystack integration, callback URLs
- Verify payment: Payment verification, participant creation, email sending
- ExcelParticipantViewSet: List/retrieve, filtering
- Webhook: Signature verification, idempotency, concurrent handling
- Security: Authentication, authorization, input validation
"""
from decimal import Decimal
from django.test import TestCase
from django.contrib.auth import get_user_model
from django.conf import settings
from django.core.files.uploadedfile import SimpleUploadedFile
from rest_framework.test import APIClient, APITestCase
from rest_framework import status
from unittest.mock import patch, Mock, MagicMock, call
from io import BytesIO
import json
import openpyxl
import hashlib
import hmac
from django.urls import reverse
from excel_bulk_orders.models import ExcelBulkOrder, ExcelCouponCode, ExcelParticipant
from excel_bulk_orders.views import excel_bulk_order_payment_webhook

User = get_user_model()


class ExcelBulkOrderViewSetTest(APITestCase):
    """Test ExcelBulkOrderViewSet CRUD operations"""

    def setUp(self):
        """Set up test data"""
        self.client = APIClient()

        # Create users
        self.regular_user = User.objects.create_user(
            username='regular',
            email='regular@example.com',
            password='testpass123'
        )

        self.staff_user = User.objects.create_user(
            username='staff',
            email='staff@example.com',
            password='testpass123',
            is_staff=True
        )

        # Create test bulk orders
        self.user_bulk_order = ExcelBulkOrder.objects.create(
            title='User Order',
            coordinator_name='User Coordinator',
            coordinator_email='usercoord@example.com',
            coordinator_phone='08012345678',
            price_per_participant=Decimal('5000.00'),
            created_by=self.regular_user
        )

        self.other_bulk_order = ExcelBulkOrder.objects.create(
            title='Other Order',
            coordinator_name='Other Coordinator',
            coordinator_email='othercoord@example.com',
            coordinator_phone='08011111111',
            price_per_participant=Decimal('6000.00')
        )

    def test_list_bulk_orders_unauthenticated(self):
        """Test listing bulk orders without authentication"""
        url = '/api/excel-bulk-orders/'
        response = self.client.get(url)

        # Should return empty for unauthenticated users
        self.assertEqual(response.status_code, status.HTTP_200_OK)

    def test_list_bulk_orders_authenticated_regular_user(self):
        """Test listing bulk orders as authenticated regular user"""
        self.client.force_authenticate(user=self.regular_user)

        url = '/api/excel-bulk-orders/'
        response = self.client.get(url)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        
        # Should only see their own bulk orders
        order_ids = [order['id'] for order in response.data['results']]
        self.assertIn(str(self.user_bulk_order.id), order_ids)
        self.assertNotIn(str(self.other_bulk_order.id), order_ids)

    def test_list_bulk_orders_authenticated_staff_user(self):
        """Test listing bulk orders as staff user"""
        self.client.force_authenticate(user=self.staff_user)

        url = '/api/excel-bulk-orders/'
        response = self.client.get(url)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        
        # Staff should see all bulk orders
        order_ids = [order['id'] for order in response.data['results']]
        self.assertIn(str(self.user_bulk_order.id), order_ids)
        self.assertIn(str(self.other_bulk_order.id), order_ids)

    @patch('excel_bulk_orders.views.generate_excel_template')
    @patch('excel_bulk_orders.views.cloudinary.uploader.upload')
    def test_create_bulk_order_success(self, mock_upload, mock_generate_template):
        """Test creating bulk order with template generation"""
        # Mock template generation
        mock_buffer = BytesIO(b'fake excel content')
        mock_generate_template.return_value = mock_buffer

        # Mock Cloudinary upload
        mock_upload.return_value = {
            'secure_url': 'https://res.cloudinary.com/test/template.xlsx'
        }

        data = {
            'title': 'New Bulk Order',
            'coordinator_name': 'New Coordinator',
            'coordinator_email': 'new@example.com',
            'coordinator_phone': '08012345678',
            'price_per_participant': '5000.00',
            'requires_custom_name': True,
        }

        url = '/api/excel-bulk-orders/'
        response = self.client.post(url, data, format='json')

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(response.data['title'], 'New Bulk Order')
        self.assertIn('template_file', response.data)
        self.assertEqual(
            response.data['template_file'],
            'https://res.cloudinary.com/test/template.xlsx'
        )

        # Verify template generation was called
        mock_generate_template.assert_called_once()
        mock_upload.assert_called_once()

    @patch('excel_bulk_orders.views.generate_excel_template')
    @patch('excel_bulk_orders.views.cloudinary.uploader.upload')
    def test_create_bulk_order_with_authenticated_user(self, mock_upload, mock_generate_template):
        """Test that created_by is set when user is authenticated"""
        self.client.force_authenticate(user=self.regular_user)

        mock_buffer = BytesIO(b'fake excel content')
        mock_generate_template.return_value = mock_buffer

        mock_upload.return_value = {
            'secure_url': 'https://res.cloudinary.com/test/template.xlsx'
        }

        data = {
            'title': 'Auth User Order',
            'coordinator_name': 'Test',
            'coordinator_email': 'authtest@example.com',
            'coordinator_phone': '08012345678',
            'price_per_participant': '5000.00',
        }

        url = '/api/excel-bulk-orders/'
        response = self.client.post(url, data, format='json')

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        # Field 'created_by' not in serializer output
        # self.assertEqual(response.data['created_by'], str(self.regular_user.id))

    @patch('excel_bulk_orders.views.generate_excel_template')
    def test_create_bulk_order_template_generation_failure(self, mock_generate_template):
        """Test handling of template generation failure"""
        # Mock template generation to raise exception
        mock_generate_template.side_effect = Exception('Template generation failed')

        data = {
            'title': 'Failed Order',
            'coordinator_name': 'Test',
            'coordinator_email': 'failed@example.com',
            'coordinator_phone': '08012345678',
            'price_per_participant': '5000.00',
        }

        url = '/api/excel-bulk-orders/'
        response = self.client.post(url, data, format='json')

        self.assertEqual(response.status_code, status.HTTP_500_INTERNAL_SERVER_ERROR)
        self.assertIn('error', response.data)

        # Bulk order should not exist in database
        self.assertFalse(
            ExcelBulkOrder.objects.filter(coordinator_email='failed@example.com').exists()
        )

    def test_retrieve_bulk_order_public_access(self):
        """Test retrieving bulk order without authentication"""
        url = f'/api/excel-bulk-orders/{self.user_bulk_order.id}/'
        response = self.client.get(url)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data['id'], str(self.user_bulk_order.id))

    def test_retrieve_nonexistent_bulk_order(self):
        """Test retrieving non-existent bulk order"""
        url = '/api/excel-bulk-orders/00000000-0000-0000-0000-000000000000/'
        response = self.client.get(url)

        self.assertEqual(response.status_code, status.HTTP_404_NOT_FOUND)

    def test_create_bulk_order_missing_required_fields(self):
        """Test creating bulk order with missing required fields"""
        data = {
            'title': 'Incomplete Order',
            # Missing coordinator_name, email, phone, price
        }

        url = '/api/excel-bulk-orders/'
        response = self.client.post(url, data, format='json')

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn('coordinator_name', response.data)
        self.assertIn('coordinator_email', response.data)
        self.assertIn('coordinator_phone', response.data)
        self.assertIn('price_per_participant', response.data)

    def test_create_bulk_order_invalid_price(self):
        """Test creating bulk order with invalid price"""
        data = {
            'title': 'Invalid Price Order',
            'coordinator_name': 'Test',
            'coordinator_email': 'test@example.com',
            'coordinator_phone': '08012345678',
            'price_per_participant': '-1000.00',  # Negative price
        }

        url = '/api/excel-bulk-orders/'
        response = self.client.post(url, data, format='json')

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)


class ExcelUploadActionTest(APITestCase):
    """Test upload Excel file action"""

    def setUp(self):
        """Set up test data"""
        self.client = APIClient()

        self.bulk_order = ExcelBulkOrder.objects.create(
            title='Upload Test Order',
            coordinator_name='Test',
            coordinator_email='upload@example.com',
            coordinator_phone='08012345678',
            price_per_participant=Decimal('5000.00'),
            template_file='https://res.cloudinary.com/test/template.xlsx'
        )

    def create_simple_excel_file(self, filename='test.xlsx'):
        """Helper to create a simple Excel file"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Participants'

        # Header
        ws.append(['S/N', 'Full Name', 'Size', 'Coupon Code'])

        # Sample data
        ws.append([1, 'John Doe', 'Medium', ''])

        # Save to BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return SimpleUploadedFile(
            filename,
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    @patch('excel_bulk_orders.views.cloudinary.uploader.upload')
    def test_upload_excel_success(self, mock_upload):
        """Test successful Excel file upload"""
        mock_upload.return_value = {
            'secure_url': 'https://res.cloudinary.com/test/upload.xlsx'
        }

        excel_file = self.create_simple_excel_file()

        url = f'/api/excel-bulk-orders/{self.bulk_order.id}/upload/'
        response = self.client.post(
            url,
            {'excel_file': excel_file},
            format='multipart'
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data['validation_status'], 'uploaded')
        self.assertEqual(
            response.data['uploaded_file'],
            'https://res.cloudinary.com/test/upload.xlsx'
        )

        # Verify bulk order was updated
        self.bulk_order.refresh_from_db()
        self.assertEqual(self.bulk_order.validation_status, 'uploaded')
        self.assertEqual(
            self.bulk_order.uploaded_file,
            'https://res.cloudinary.com/test/upload.xlsx'
        )

    def test_upload_excel_wrong_extension(self):
        """Test uploading file with wrong extension"""
        csv_file = SimpleUploadedFile(
            'test.csv',
            b'name,size\nJohn,M',
            content_type='text/csv'
        )

        url = f'/api/excel-bulk-orders/{self.bulk_order.id}/upload/'
        response = self.client.post(
            url,
            {'excel_file': csv_file},
            format='multipart'
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn('excel_file', response.data)

    def test_upload_excel_file_too_large(self):
        """Test uploading file exceeding size limit"""
        # Create a large file (> 5MB)
        large_content = b'X' * (6 * 1024 * 1024)
        large_file = SimpleUploadedFile(
            'large.xlsx',
            large_content,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        url = f'/api/excel-bulk-orders/{self.bulk_order.id}/upload/'
        response = self.client.post(
            url,
            {'excel_file': large_file},
            format='multipart'
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn('excel_file', response.data)

    def test_upload_excel_already_paid(self):
        """Test uploading Excel when payment is already completed"""
        self.bulk_order.payment_status = True
        self.bulk_order.save()

        excel_file = self.create_simple_excel_file()

        url = f'/api/excel-bulk-orders/{self.bulk_order.id}/upload/'
        response = self.client.post(
            url,
            {'excel_file': excel_file},
            format='multipart'
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn('error', response.data)
        self.assertIn('already completed', response.data['error'])

    def test_upload_excel_missing_file(self):
        """Test upload without providing file"""
        url = f'/api/excel-bulk-orders/{self.bulk_order.id}/upload/'
        response = self.client.post(url, {}, format='multipart')

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn('excel_file', response.data)

    @patch('excel_bulk_orders.views.cloudinary.uploader.upload')
    def test_upload_excel_cloudinary_failure(self, mock_upload):
        """Test handling of Cloudinary upload failure"""
        mock_upload.side_effect = Exception('Cloudinary upload failed')

        excel_file = self.create_simple_excel_file()

        url = f'/api/excel-bulk-orders/{self.bulk_order.id}/upload/'
        response = self.client.post(
            url,
            {'excel_file': excel_file},
            format='multipart'
        )

        self.assertEqual(response.status_code, status.HTTP_500_INTERNAL_SERVER_ERROR)
        self.assertIn('error', response.data)

    def test_upload_to_nonexistent_bulk_order(self):
        """Test uploading to non-existent bulk order"""
        excel_file = self.create_simple_excel_file()

        url = '/api/excel-bulk-orders/00000000-0000-0000-0000-000000000000/upload/'
        response = self.client.post(
            url,
            {'excel_file': excel_file},
            format='multipart'
        )

        self.assertEqual(response.status_code, status.HTTP_404_NOT_FOUND)


class ExcelValidateActionTest(APITestCase):
    """Test validate Excel file action"""

    def setUp(self):
        """Set up test data"""
        self.client = APIClient()

        self.bulk_order = ExcelBulkOrder.objects.create(
            title='Validate Test Order',
            coordinator_name='Test',
            coordinator_email='validate@example.com',
            coordinator_phone='08012345678',
            price_per_participant=Decimal('5000.00'),
            template_file='https://res.cloudinary.com/test/template.xlsx',
            uploaded_file='https://res.cloudinary.com/test/upload.xlsx',
            validation_status='uploaded'
        )

    @staticmethod
    def create_valid_excel_bytes():
        """Create valid Excel file bytes for mocking"""
        import openpyxl
        from io import BytesIO
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Participants"
        ws['A1'], ws['B1'], ws['C1'], ws['D1'] = 'S/N', 'Full Name', 'Size', 'Coupon Code'
        for i in range(2, 12):  # 10 rows
            ws[f'A{i}'], ws[f'B{i}'], ws[f'C{i}'], ws[f'D{i}'] = i-1, f'Participant {i-1}', 'Medium', ''
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    @patch('excel_bulk_orders.views.validate_excel_file')
    @patch('requests.get')
    def test_validate_excel_success(self, mock_requests_get, mock_validate):
        """Test successful Excel validation"""
        # Mock file download with valid Excel bytes
        mock_response = Mock()
        mock_response.content = self.create_valid_excel_bytes()
        mock_requests_get.return_value = mock_response

        # Mock validation
        mock_validate.return_value = {
            'valid': True,
            'errors': [],
            'summary': {
                'total_rows': 10,
                'valid_rows': 10,
                'error_rows': 0
            }
        }

        url = f'/api/excel-bulk-orders/{self.bulk_order.id}/validate/'
        response = self.client.post(url)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data['validation_result']['valid'])
        self.assertEqual(response.data['bulk_order']['validation_status'], 'valid')

        # Verify bulk order was updated
        self.bulk_order.refresh_from_db()
        self.assertEqual(self.bulk_order.validation_status, 'valid')
        # Total amount should be calculated: 10 participants * 5000
        self.assertEqual(self.bulk_order.total_amount, Decimal('50000.00'))

    @patch('excel_bulk_orders.views.validate_excel_file')
    @patch('requests.get')
    def test_validate_excel_with_errors(self, mock_requests_get, mock_validate):
        """Test Excel validation with errors"""
        # Mock file download
        mock_response = Mock()
        mock_response.content = b'fake excel content'
        mock_requests_get.return_value = mock_response

        # Mock validation with errors
        mock_validate.return_value = {
            'valid': False,
            'errors': [
                {
                    'row': 3,
                    'field': 'Size',
                    'error': 'Invalid size value',
                    'current_value': 'XL+'
                }
            ],
            'summary': {
                'total_rows': 5,
                'valid_rows': 4,
                'error_rows': 1
            }
        }

        url = f'/api/excel-bulk-orders/{self.bulk_order.id}/validate/'
        response = self.client.post(url)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertFalse(response.data['validation_result']['valid'])
        self.assertEqual(response.data['bulk_order']['validation_status'], 'invalid')
        self.assertEqual(len(response.data['validation_result']['errors']), 1)

        # Verify bulk order was updated
        self.bulk_order.refresh_from_db()
        self.assertEqual(self.bulk_order.validation_status, 'invalid')
        self.assertEqual(self.bulk_order.total_amount, Decimal('0.00'))

    @patch('excel_bulk_orders.utils.validate_excel_file')
    def test_validate_excel_with_coupons(self, mock_validate):
        """Test Excel validation with some coupon entries"""
        from excel_bulk_orders.models import ExcelCouponCode  # CORRECTED
        from unittest.mock import Mock
        
        # Create coupons for this bulk order
        coupon1 = ExcelCouponCode.objects.create(
            bulk_order=self.bulk_order,
            code='TESTCOUPON1'
        )
        coupon2 = ExcelCouponCode.objects.create(
            bulk_order=self.bulk_order,
            code='TESTCOUPON2'
        )
        
        # Mock validation result
        mock_validate.return_value = {
            'valid': True,
            'errors': [],
            'summary': {
                'total_rows': 5,
                'valid_rows': 5,
                'error_rows': 0
            }
        }
        
        # Set uploaded file
        self.bulk_order.uploaded_file = 'https://example.com/uploaded.xlsx'
        self.bulk_order.save()
        
        # Mock Excel reading
        with patch('excel_bulk_orders.views.pd.read_excel') as mock_read_excel:
            # Create mock DataFrame
            mock_df = Mock()
            mock_df.__len__ = Mock(return_value=5)
            mock_df.iterrows.return_value = [
                (0, {'Full Name': 'John Doe', 'Size': 'Large', 'Coupon Code': ''}),
                (1, {'Full Name': 'Jane Smith', 'Size': 'Medium', 'Coupon Code': 'TESTCOUPON1'}),
                (2, {'Full Name': 'Bob Johnson', 'Size': 'Small', 'Coupon Code': ''}),
                (3, {'Full Name': 'Alice Brown', 'Size': 'Large', 'Coupon Code': 'TESTCOUPON2'}),
                (4, {'Full Name': 'Charlie Wilson', 'Size': 'Medium', 'Coupon Code': ''}),
            ]
            mock_read_excel.return_value = mock_df
            
            # Call validate endpoint
            url = reverse('excelbulkorder-validate', kwargs={'pk': self.bulk_order.pk})
            response = self.client.post(url)
        
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.data['is_valid'])
        self.assertEqual(response.data['total_rows'], 5)
        
        # Refresh bulk order
        self.bulk_order.refresh_from_db()
        
        # CORRECTED: 5 total - 2 with valid coupons = 3 chargeable Ã— 10,000 = 30,000
        expected_amount = Decimal('30000.00')
        self.assertEqual(self.bulk_order.total_amount, expected_amount)
        
        # Should be marked as validated
        self.assertEqual(self.bulk_order.validation_status, 'valid')



    def test_validate_excel_not_uploaded(self):
        """Test validation when Excel not uploaded yet"""
        bulk_order = ExcelBulkOrder.objects.create(
            title='Not Uploaded Order',
            coordinator_name='Test',
            coordinator_email='notupload@example.com',
            coordinator_phone='08012345678',
            price_per_participant=Decimal('5000.00'),
            validation_status='pending'
        )

        url = f'/api/excel-bulk-orders/{bulk_order.id}/validate/'
        response = self.client.post(url)

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn('error', response.data)

    def test_validate_excel_already_paid(self):
        """Test validation when payment is already completed"""
        self.bulk_order.payment_status = True
        self.bulk_order.save()

        url = f'/api/excel-bulk-orders/{self.bulk_order.id}/validate/'
        response = self.client.post(url)

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn('error', response.data)

    @patch('requests.get')
    def test_validate_excel_file_download_failure(self, mock_requests_get):
        """Test handling of file download failure"""
        mock_requests_get.side_effect = Exception('Download failed')

        url = f'/api/excel-bulk-orders/{self.bulk_order.id}/validate/'
        response = self.client.post(url)

        self.assertEqual(response.status_code, status.HTTP_500_INTERNAL_SERVER_ERROR)
        self.assertIn('error', response.data)


class ExcelInitializePaymentActionTest(APITestCase):
    """Test initialize payment action"""

    def setUp(self):
        """Set up test data"""
        self.client = APIClient()

        self.bulk_order = ExcelBulkOrder.objects.create(
            title='Payment Test Order',
            coordinator_name='Test Coordinator',
            coordinator_email='payment@example.com',
            coordinator_phone='08012345678',
            price_per_participant=Decimal('5000.00'),
            total_amount=Decimal('25000.00'),
            validation_status='valid',
            uploaded_file='https://res.cloudinary.com/test/upload.xlsx'
        )

    @patch('excel_bulk_orders.views.initialize_payment')
    def test_initialize_payment_success(self, mock_initialize):
        """Test successful payment initialization"""
        mock_initialize.return_value = {
            'status': True,
            'data': {
                'authorization_url': 'https://checkout.paystack.com/test',
                'access_code': 'test_access_code',
                'reference': self.bulk_order.reference
            }
        }

        url = f'/api/excel-bulk-orders/{self.bulk_order.id}/initialize-payment/'
        response = self.client.post(url)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn('authorization_url', response.data)
        self.assertIn('access_code', response.data)
        self.assertIn('reference', response.data)

        # Verify bulk order status updated
        self.bulk_order.refresh_from_db()
        self.assertEqual(self.bulk_order.validation_status, 'processing')

    def test_initialize_payment_not_validated(self):
        """Test payment initialization when Excel not validated"""
        self.bulk_order.validation_status = 'uploaded'
        self.bulk_order.save()

        url = f'/api/excel-bulk-orders/{self.bulk_order.id}/initialize-payment/'
        response = self.client.post(url)

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn('error', response.data)
        self.assertIn('validated', response.data['error'])

    def test_initialize_payment_already_paid(self):
        """Test payment initialization when already paid"""
        self.bulk_order.payment_status = True
        self.bulk_order.save()

        url = f'/api/excel-bulk-orders/{self.bulk_order.id}/initialize-payment/'
        response = self.client.post(url)

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn('error', response.data)
        self.assertIn('already completed', response.data['error'])

    def test_initialize_payment_zero_amount(self):
        """Test payment initialization with zero amount"""
        self.bulk_order.total_amount = Decimal('0.00')
        self.bulk_order.save()

        url = f'/api/excel-bulk-orders/{self.bulk_order.id}/initialize-payment/'
        response = self.client.post(url)

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn('error', response.data)
        self.assertIn('Invalid payment amount', response.data['error'])

    @patch('excel_bulk_orders.views.initialize_payment')
    def test_initialize_payment_paystack_failure(self, mock_initialize):
        """Test handling of Paystack initialization failure"""
        mock_initialize.return_value = {
            'status': False
        }

        url = f'/api/excel-bulk-orders/{self.bulk_order.id}/initialize-payment/'
        response = self.client.post(url)

        self.assertEqual(response.status_code, status.HTTP_500_INTERNAL_SERVER_ERROR)
        self.assertIn('error', response.data)

    @patch('excel_bulk_orders.views.initialize_payment')
    def test_initialize_payment_callback_url(self, mock_initialize):
        """Test that callback URL is properly constructed"""
        mock_initialize.return_value = {
            'status': True,
            'data': {
                'authorization_url': 'https://checkout.paystack.com/test',
                'access_code': 'test_access_code',
                'reference': self.bulk_order.reference
            }
        }

        url = f'/api/excel-bulk-orders/{self.bulk_order.id}/initialize-payment/'
        response = self.client.post(url)

        self.assertEqual(response.status_code, status.HTTP_200_OK)

        # Verify initialize_payment was called with callback_url
        call_args = mock_initialize.call_args
        self.assertIn('callback_url', call_args[1])
        self.assertIn('verify-payment', call_args[1]['callback_url'])


class ExcelVerifyPaymentActionTest(APITestCase):
    """Test verify payment action"""

    def setUp(self):
        """Set up test data"""
        self.client = APIClient()

        self.bulk_order = ExcelBulkOrder.objects.create(
            title='Verify Payment Order',
            coordinator_name='Test Coordinator',
            coordinator_email='verify@example.com',
            coordinator_phone='08012345678',
            price_per_participant=Decimal('5000.00'),
            total_amount=Decimal('25000.00'),
            validation_status='processing',
            uploaded_file='https://res.cloudinary.com/test/upload.xlsx'
        )

        # Create coupon
        self.coupon = ExcelCouponCode.objects.create(
            bulk_order=self.bulk_order,
            code='TEST123'
        )

    def create_test_excel_file(self):
        """Helper to create a test Excel file with participants"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Participants'

        # Header
        ws.append(['S/N', 'Full Name', 'Size', 'Coupon Code'])

        # Participants
        ws.append([1, 'John Doe', 'Medium', ''])
        ws.append([2, 'Jane Smith', 'Large', ''])
        ws.append([3, 'Bob Johnson', 'Small', 'TEST123'])

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return buffer.getvalue()

    @patch('excel_bulk_orders.views.send_bulk_order_confirmation_email')
    @patch('excel_bulk_orders.views.create_participants_from_excel')
    @patch('requests.get')
    @patch('excel_bulk_orders.views.verify_payment')
    def test_verify_payment_success(
        self, mock_verify, mock_requests_get, mock_create_participants, mock_send_email
    ):
        """Test successful payment verification and participant creation"""
        # Mock payment verification
        mock_verify.return_value = {
            'status': True,
            'data': {
                'status': 'success',
                'amount': 2500000,  # 25000.00 in kobo
                'reference': self.bulk_order.reference
            }
        }

        # Mock Excel file download
        mock_response = Mock()
        mock_response.content = self.create_test_excel_file()
        mock_requests_get.return_value = mock_response

        # Mock participant creation
        mock_create_participants.return_value = 3

        url = f'/api/excel-bulk-orders/{self.bulk_order.id}/verify-payment/'
        response = self.client.post(url)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn('message', response.data)

        # Verify bulk order was updated
        self.bulk_order.refresh_from_db()
        self.assertTrue(self.bulk_order.payment_status)
        self.assertEqual(self.bulk_order.validation_status, 'completed')
        self.assertEqual(self.bulk_order.paystack_reference, self.bulk_order.reference)

        # Verify participants were created
        mock_create_participants.assert_called_once()

        # Email is sent asynchronously via background task, not directly
        # So we can't assert on the mock - it's queued for later processing
        # mock_send_email.assert_called_once()

    @patch('excel_bulk_orders.views.verify_payment')
    def test_verify_payment_already_verified(self, mock_verify):
        """Test verifying payment that's already verified"""
        self.bulk_order.payment_status = True
        self.bulk_order.save()

        url = f'/api/excel-bulk-orders/{self.bulk_order.id}/verify-payment/'
        response = self.client.post(url)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn('already verified', response.data['message'])

        # verify_payment should not be called
        mock_verify.assert_not_called()

    @patch('excel_bulk_orders.views.verify_payment')
    def test_verify_payment_failed_verification(self, mock_verify):
        """Test handling of failed payment verification"""
        mock_verify.return_value = {
            'status': False
        }

        url = f'/api/excel-bulk-orders/{self.bulk_order.id}/verify-payment/'
        response = self.client.post(url)

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn('error', response.data)

    @patch('excel_bulk_orders.views.verify_payment')
    def test_verify_payment_unsuccessful_status(self, mock_verify):
        """Test handling of unsuccessful payment status"""
        mock_verify.return_value = {
            'status': True,
            'data': {
                'status': 'failed',
                'reference': self.bulk_order.reference
            }
        }

        url = f'/api/excel-bulk-orders/{self.bulk_order.id}/verify-payment/'
        response = self.client.post(url)

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn('error', response.data)

    @patch('excel_bulk_orders.views.create_participants_from_excel')
    @patch('requests.get')
    @patch('excel_bulk_orders.views.verify_payment')
    def test_verify_payment_participant_creation_failure(
        self, mock_verify, mock_requests_get, mock_create_participants
    ):
        """Test handling of participant creation failure"""
        # Mock successful payment verification
        mock_verify.return_value = {
            'status': True,
            'data': {
                'status': 'success',
                'reference': self.bulk_order.reference
            }
        }

        # Mock file download
        mock_response = Mock()
        mock_response.content = self.create_test_excel_file()
        mock_requests_get.return_value = mock_response

        # Mock participant creation failure
        mock_create_participants.side_effect = Exception('Participant creation failed')

        url = f'/api/excel-bulk-orders/{self.bulk_order.id}/verify-payment/'
        response = self.client.post(url)

        self.assertEqual(response.status_code, status.HTTP_500_INTERNAL_SERVER_ERROR)
        self.assertIn('error', response.data)

    @patch('excel_bulk_orders.views.verify_payment')
    def test_verify_payment_with_reference_parameter(self, mock_verify):
        """Test verify payment with reference in query params"""
        mock_verify.return_value = {
            'status': False
        }

        custom_reference = 'CUSTOM-REF-123'
        url = f'/api/excel-bulk-orders/{self.bulk_order.id}/verify-payment/?reference={custom_reference}'
        response = self.client.post(url)

        # Verify that custom reference was used
        mock_verify.assert_called_once_with(custom_reference)

    def test_verify_payment_get_method_allowed(self):
        """Test that GET method is also allowed for verify payment"""
        url = f'/api/excel-bulk-orders/{self.bulk_order.id}/verify-payment/'
        response = self.client.get(url)

        # Should not return 405 Method Not Allowed
        self.assertNotEqual(response.status_code, status.HTTP_405_METHOD_NOT_ALLOWED)